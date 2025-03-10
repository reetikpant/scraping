import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def fetch_and_parse_hotels(url):
    try:
        # Fetch the webpage
        response = requests.get(url)
        response.raise_for_status()
        
        # Parse the HTML content
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Initialize lists to store data
        data = []
        
        # Find all city sections
        cities = soup.find_all('b', class_='city_01')
        
        for city in cities:
            city_name = city.find('font').text
            
            # Find the next unordered list that contains hotels
            hotel_list = city.find_next('ul')
            
            # Process each hotel in the list
            for hotel in hotel_list.find_all('li'):
                hotel_info = {}
                hotel_info['City'] = city_name
                
                # Get hotel name with a safety check
                hotel_name_tag = hotel.find('b', class_='strong_001')
                if hotel_name_tag:
                    hotel_info['Hotel Name'] = hotel_name_tag.text
                else:
                    hotel_info['Hotel Name'] = 'N/A'  # Assign default value if name is not found
                
                # Get all text and split into lines
                lines = [line.strip() for line in hotel.get_text().split('\n') if line.strip()]
                
                # Remove hotel name from lines as we already have it
                lines = lines[1:]
                
                # Process each line to extract information
                for line in lines:
                    if 'Tel' in line:
                        hotel_info['Telephone'] = line.replace('Tel :', '').strip()
                    elif 'Fax' in line:
                        hotel_info['Fax'] = line.replace('Fax :', '').strip()
                    elif 'Email' in line.lower() or '@' in line:
                        hotel_info['Email'] = line.split(':')[-1].strip()
                    elif 'Web' in line:
                        hotel_info['Website'] = line.replace('Web :', '').strip()
                    elif '@' in line:
                        hotel_info['Email'] = line.strip()
                    elif 'www.' in line.lower() or 'http' in line.lower():
                        hotel_info['Website'] = line.strip()
                    else:
                        if 'Address' not in hotel_info:
                            hotel_info['Address'] = line
                        else:
                            hotel_info['Address'] += ', ' + line
                
                data.append(hotel_info)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Reorder columns
        column_order = ['City', 'Hotel Name', 'Address', 'Telephone', 'Fax', 'Email', 'Website']
        df = df.reindex(columns=column_order)
        
        return df
    
    except requests.exceptions.RequestException as e:
        print(f"Error fetching the webpage: {e}")
        return None


def save_to_excel(df, filename=None):
    if df is None:
        return
    
    if filename is None:
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f'hotels_data_{current_date}.xlsx'
    
    # Create Excel writer object with openpyxl engine
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, sheet_name='Hotels', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Hotels']
        
        # Define formats using openpyxl
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15  # City
        worksheet.column_dimensions['B'].width = 25  # Hotel Name
        worksheet.column_dimensions['C'].width = 40  # Address
        worksheet.column_dimensions['D'].width = 20  # Telephone
        worksheet.column_dimensions['E'].width = 20  # Fax
        worksheet.column_dimensions['F'].width = 30  # Email
        worksheet.column_dimensions['G'].width = 35  # Website
        
        # Apply header formatting
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border
        
        # Apply cell formatting to all data cells
        for row_num in range(len(df)):
            for col_num in range(len(df.columns)):
                cell = worksheet.cell(row=row_num + 2, column=col_num + 1, value=df.iloc[row_num, col_num])
                cell.alignment = alignment
                cell.border = border
    
    print(f"Data successfully saved to {filename}")



def fetch_and_parse_hotels_for_multiple_urls(base_url, pages):
    all_data = []

    for page in pages:
        url = f"{base_url}/{page}.html"  # Construct the URL (e.g., 'a.html', 'b.html', ...)
        print(f"Processing {url}...")
        df = fetch_and_parse_hotels(url)  # Fetch and parse data from the current page
        if df is not None:
            all_data.append(df)

    # Concatenate all DataFrames into a single DataFrame
    if all_data:
        final_df = pd.concat(all_data, ignore_index=True)
        return final_df
    return None

# Main execution
base_url = "https://hotelassociationofindia.com"  # Example base URL
pages = ['a', 'b', 'c', 'd', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'r', 's', 't', 'u', 'v']  # List of page identifiers
final_df = fetch_and_parse_hotels_for_multiple_urls(base_url, pages)

if final_df is not None:
    save_to_excel(final_df)  # Save the combined data to Excel

