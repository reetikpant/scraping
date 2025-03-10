import os
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# Set up Selenium WebDriver
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# File paths
excel_file = "scraped_data_with_selenium.xlsx"  # Replace with your Excel file name
image_dir = "images"  # Directory to save images

# Create the image directory if it doesn't exist
os.makedirs(image_dir, exist_ok=True)

# Open the Excel file
wb = load_workbook(excel_file)
ws = wb.active

# Headers to mimic a real browser
headers = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/109.0.5414.119 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.5",
}

# Iterate through each row (assuming the headers are in the first row)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    author_cell = row[2]  # Assuming Author is in column C (index 2)
    image_url_cell = row[0]  # Assuming Image URL is in column A (index 0)
    title_cell = row[1]  # Assuming Title is in column B (index 1)
    email_cell = row[4]  # Assuming Email column is in column E (index 4)

    author = author_cell.value
    image_url = image_url_cell.value
    title = title_cell.value

    # Fetch email using Selenium
    try:
        url = f"https://photo.ntb.gov.np/{author}"
        driver.get(url)

        email_element = driver.find_element(By.CLASS_NAME, 'email')
        email = email_element.get_attribute('href').replace('mailto:', '')
        email_cell.value = email  # Save the email in the Excel file
        print(f"Email for {author}: {email}")
    except Exception as e:
        print(f"Failed to fetch email for {author}: {e}")
        email_cell.value = "Email not found"

    # Download the image
    try:
        response = requests.get(image_url, stream=True, headers=headers)
        if response.status_code == 200:
            # Save the image with the title as the filename
            filename = os.path.join(image_dir, f"{title}.jpg")
            with open(filename, 'wb') as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)
            print(f"High-quality image saved: {filename}")
        else:
            print(f"Failed to download high-quality image for {title}")
    except Exception as e:
        print(f"Error downloading image for {title}: {e}")

# Save the updated Excel file
wb.save(excel_file)

# Close the browser
driver.quit()
